#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
売上データ分析ChatGPT研修用Excelファイル作成スクリプト
"""

import os
import random
from datetime import datetime, timedelta
from openpyxl import Workbook

# 商品データの定義
PRODUCTS = [
    {'name': 'ノートPC', 'category': 'PC', 'base_price': 80000},
    {'name': 'デスクトップPC', 'category': 'PC', 'base_price': 120000},
    {'name': 'タブレット', 'category': 'PC', 'base_price': 40000},
    {'name': 'スマートフォン', 'category': 'モバイル', 'base_price': 60000},
    {'name': 'イヤホン', 'category': 'アクセサリ', 'base_price': 8000},
    {'name': 'マウス', 'category': 'アクセサリ', 'base_price': 3000},
    {'name': 'キーボード', 'category': 'アクセサリ', 'base_price': 12000},
    {'name': 'モニター', 'category': 'PC', 'base_price': 25000},
    {'name': 'プリンター', 'category': 'その他', 'base_price': 15000},
    {'name': 'Webカメラ', 'category': 'アクセサリ', 'base_price': 5000},
]

REGIONS = ['東京', '大阪', '名古屋', '福岡', '札幌']

def create_sales_data_monthly():
    """動画1用：sales_data_monthly.xlsx"""
    wb = Workbook()
    ws = wb.active
    ws.title = '売上データ'
    
    # ヘッダー設定
    headers = ['日付', '商品名', 'カテゴリ', '単価', '数量', '売上金額', '地域']
    ws.append(headers)
    
    # 2024年のデータ生成（約50行）
    current_date = datetime(2024, 1, 1)
    data = []
    
    for _ in range(50):
        # 月をランダムに進める
        month_add = random.randint(0, 11)
        date = current_date.replace(month=((current_date.month + month_add - 1) % 12) + 1)
        
        # ランダムな商品選択
        product = random.choice(PRODUCTS)
        quantity = random.randint(1, 10)
        # 価格に若干の変動を追加
        price = product['base_price'] + random.randint(-5000, 5000)
        sales_amount = price * quantity
        region = random.choice(REGIONS)
        
        row = [
            date.strftime('%Y/%m/%d'),
            product['name'],
            product['category'],
            price,
            quantity,
            sales_amount,
            region
        ]
        ws.append(row)
        data.append(row)
    
    return wb

def create_sales_for_graph():
    """動画2用：sales_for_graph.xlsx"""
    wb = Workbook()
    
    # シート1：月次集計
    ws1 = wb.active
    ws1.title = '月次集計'
    
    headers1 = ['月', '2024年売上', '2023年売上', '前年比']
    ws1.append(headers1)
    
    for month in range(1, 13):
        sales_2024 = random.randint(500000, 2000000)
        sales_2023 = random.randint(400000, 1800000)
        ratio = round((sales_2024 / sales_2023) * 100, 1)
        
        ws1.append([f'{month}月', sales_2024, sales_2023, f'{ratio}%'])
    
    # シート2：商品別
    ws2 = wb.create_sheet('商品別')
    headers2 = ['商品カテゴリ', '売上金額', '構成比']
    ws2.append(headers2)
    
    categories = list(set(p['category'] for p in PRODUCTS))
    total_sales = 0
    category_sales = {}
    
    for category in categories:
        sales = random.randint(1000000, 5000000)
        category_sales[category] = sales
        total_sales += sales
    
    for category, sales in category_sales.items():
        ratio = round((sales / total_sales) * 100, 1)
        ws2.append([category, sales, f'{ratio}%'])
    
    return wb

def create_sales_history_2years():
    """動画3用：sales_history_2years.xlsx"""
    wb = Workbook()
    ws = wb.active
    ws.title = '過去2年分'
    
    headers = ['年月', '売上金額', '前年同月', '季節係数']
    ws.append(headers)
    
    # 季節性を考慮したデータ生成
    seasonal_factors = {
        1: 0.8, 2: 0.7, 3: 1.0, 4: 0.9, 5: 0.9, 6: 1.0,
        7: 1.1, 8: 1.2, 9: 1.0, 10: 1.0, 11: 1.3, 12: 1.5
    }
    
    base_sales = 1000000
    
    # 2022-2024年のデータ
    for year in [2022, 2023, 2024]:
        for month in range(1, 13):
            seasonal_factor = seasonal_factors[month]
            # 年による成長率
            growth_factor = 1.0 + (year - 2022) * 0.1
            sales = int(base_sales * seasonal_factor * growth_factor * random.uniform(0.8, 1.2))
            
            if year > 2022:
                # 前年同月データがある場合
                prev_year_sales = int(base_sales * seasonal_factor * (growth_factor - 0.1) * random.uniform(0.8, 1.2))
            else:
                prev_year_sales = None
            
            ws.append([f'{year}/{month:02d}', sales, prev_year_sales, seasonal_factor])
    
    return wb

def create_sales_for_dashboard():
    """動画4用：sales_for_dashboard.xlsx"""
    wb = Workbook()
    ws = wb.active
    ws.title = 'ダッシュボード用'
    
    headers = ['日付', '商品ID', '商品名', 'カテゴリ', '地域', '販売員', '単価', '数量', '売上金額', '利益率']
    ws.append(headers)
    
    sales_staff = ['田中', '佐藤', '鈴木', '高橋', '渡辺', '伊藤', '山本', '中村']
    
    # より詳細なダッシュボード用データ
    for i in range(100):
        date = datetime(2024, 1, 1) + timedelta(days=random.randint(0, 365))
        product = random.choice(PRODUCTS)
        product_id = f'P{random.randint(1000, 9999)}'
        region = random.choice(REGIONS)
        staff = random.choice(sales_staff)
        quantity = random.randint(1, 5)
        price = product['base_price'] + random.randint(-3000, 3000)
        sales_amount = price * quantity
        profit_rate = round(random.uniform(15, 35), 1)
        
        ws.append([
            date.strftime('%Y-%m-%d'),
            product_id,
            product['name'],
            product['category'],
            region,
            staff,
            price,
            quantity,
            sales_amount,
            f'{profit_rate}%'
        ])
    
    return wb

def create_sales_with_errors():
    """動画5用：sales_with_errors.xlsx"""
    wb = Workbook()
    ws = wb.active
    ws.title = '問題のあるデータ'
    
    headers = ['日付', '商品名', 'カテゴリ', '単価', '数量', '売上金額', '地域', '備考']
    ws.append(headers)
    
    # 意図的にエラーのあるデータを作成
    error_data = [
        # 正常なデータ
        ['2024/01/15', 'ノートPC', 'PC', 80000, 2, 160000, '東京', '正常'],
        # 日付エラー
        ['2024/13/01', 'マウス', 'アクセサリ', 3000, 1, 3000, '大阪', '日付エラー'],
        ['不正な日付', 'キーボード', 'アクセサリ', 12000, 1, 12000, '名古屋', '日付形式エラー'],
        # 数値エラー
        ['2024/02/10', 'タブレット', 'PC', 'エラー', 1, 40000, '福岡', '価格が文字'],
        ['2024/02/15', 'スマートフォン', 'モバイル', 60000, '多数', 120000, '札幌', '数量が文字'],
        # 欠損値
        ['2024/03/01', '', 'PC', 25000, 1, 25000, '東京', '商品名欠損'],
        ['2024/03/05', 'プリンター', '', 15000, 2, 30000, '大阪', 'カテゴリ欠損'],
        ['2024/03/10', 'イヤホン', 'アクセサリ', 8000, 3, 24000, '', '地域欠損'],
        # 文字化け風
        ['2024/04/01', '?????', '???', 0, 1, 0, '???', '文字化け'],
        # 計算エラー
        ['2024/04/15', 'Webカメラ', 'アクセサリ', 5000, 2, 15000, '名古屋', '計算間違い'],
        # 重複データ
        ['2024/05/01', 'モニター', 'PC', 25000, 1, 25000, '東京', '重複1'],
        ['2024/05/01', 'モニター', 'PC', 25000, 1, 25000, '東京', '重複2'],
    ]
    
    for row in error_data:
        ws.append(row)
    
    return wb

def main():
    base_dir = "/Users/keisuke/git/AIフロー/売上データ分析ChatGPT研修"
    
    # 各動画ディレクトリとファイル名の対応
    files_to_create = [
        ("動画1_Excelに頼らずChatGPTで売上CSVを読ませる初めの一歩", "sales_data_monthly.xlsx", create_sales_data_monthly),
        ("動画2_グラフもAIにおまかせPythonコードをChatGPTに書かせる技術", "sales_for_graph.xlsx", create_sales_for_graph),
        ("動画3_来月の売上を当てろChatGPTに予測させるプロンプト術", "sales_history_2years.xlsx", create_sales_history_2years),
        ("動画4_作業ゼロで見える化LookerStudioで自動更新ダッシュボード", "sales_for_dashboard.xlsx", create_sales_for_dashboard),
        ("動画5_ChatGPTがうまく働かない指示ミスを防ぐ5つの型", "sales_with_errors.xlsx", create_sales_with_errors)
    ]
    
    for dir_name, file_name, create_func in files_to_create:
        dir_path = os.path.join(base_dir, dir_name)
        if os.path.exists(dir_path):
            file_path = os.path.join(dir_path, file_name)
            wb = create_func()
            wb.save(file_path)
            print(f"Created: {file_path}")
        else:
            print(f"Directory not found: {dir_path}")

if __name__ == "__main__":
    main()